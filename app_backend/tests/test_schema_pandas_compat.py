import asyncio
import base64
import io
import json
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
import plotly.graph_objects as go
import pytest
from openpyxl import load_workbook
from utils.analyst_db import DatasetMetadata, DatasetType, InternalDataSourceType
from utils.schema import (
    AnalystChatMessage,
    DataDictionary,
    RunAnalysisResult,
    RunAnalysisResultMetadata,
    RunChartsResult,
)


def test_data_dictionary_from_analyst_df_stringifies_tuple_columns() -> None:
    df = pd.DataFrame(
        [[100, 200]],
        columns=pd.Index(
            [
                ("black_friday", 0.0),
                ("holiday", 1.0),
            ]
        ),
    )

    dictionary = DataDictionary.from_analyst_df(df)

    assert [column.column for column in dictionary.column_descriptions] == [
        "('black_friday', 0.0)",
        "('holiday', 1.0)",
    ]
    assert [column.data_type for column in dictionary.column_descriptions] == [
        "int64",
        "int64",
    ]


def test_analysis_result_download_uses_pandas_csv_writer() -> None:
    asyncio.run(_assert_analysis_result_download_uses_pandas_csv_writer())


async def _assert_analysis_result_download_uses_pandas_csv_writer() -> None:
    from core.routers.datasets import download_dataset

    analyst_db = _FakeAnalystDB()

    response = await download_dataset(
        dataset_id="dataset-123456789",
        analyst_db=analyst_db,  # type: ignore[arg-type]
        bom=True,
    )

    csv_text = response.body.decode("utf-8")
    assert csv_text == "\ufeffamount,label\n10,A\n20,B\n"


def test_chat_excel_export_writes_pandas_analysis_result_sheet() -> None:
    asyncio.run(_assert_chat_excel_export_writes_pandas_analysis_result_sheet())


def test_chat_excel_export_writes_problematic_plotly_trace_sheet(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    asyncio.run(
        _assert_chat_excel_export_writes_problematic_plotly_trace_sheet(monkeypatch)
    )


def test_run_charts_result_uses_japanese_plotly_font_fallbacks() -> None:
    fig_json = go.Figure(data=[go.Bar(x=["東京", "大阪"], y=[1, 2])]).to_json()

    result = RunChartsResult(
        status="success",
        fig1_json=fig_json,
        fig2_json=fig_json,
        metadata=RunAnalysisResultMetadata(duration=0, attempts=1),
    )

    expected_family = "Noto Sans CJK JP, Noto Sans JP, sans-serif"
    assert result.fig1 is not None
    assert result.fig1.layout.font.family == expected_family
    assert result.fig2 is not None
    assert result.fig2.layout.font.family == expected_family


def test_plotly_trace_export_handles_ragged_arrays_without_pandas_error() -> None:
    from core.customize.infrastructure.export.chat_export import (
        plotly_trace_to_dataframe,
    )

    chart_df = plotly_trace_to_dataframe(
        {
            "type": "bar",
            "x": ["FY2025", "FY2026"],
            "y": [100],
        }
    )

    assert chart_df["x"].to_list() == ["FY2025", "FY2026"]
    assert chart_df.loc[0, "y"] == 100
    assert pd.isna(chart_df.loc[1, "y"])


def test_plotly_trace_export_ignores_nested_style_dicts_without_pandas_error() -> None:
    from core.customize.infrastructure.export.chat_export import (
        plotly_trace_to_dataframe,
    )

    chart_df = plotly_trace_to_dataframe(
        {
            "type": "scatter",
            "x": ["A", "B"],
            "y": [10, 20],
            "line": {"color": "red"},
            "marker": {"size": [4, 8]},
        }
    )

    assert chart_df.to_dict("records") == [
        {"x": "A", "y": 10},
        {"x": "B", "y": 20},
    ]
    assert "line" not in chart_df.columns
    assert "marker" not in chart_df.columns


async def _assert_chat_excel_export_writes_pandas_analysis_result_sheet() -> None:
    from core.routers.chats import save_chat_messages

    analyst_db = _FakeAnalystDB()

    response = await save_chat_messages(
        request=None,  # type: ignore[arg-type]
        chat_id="chat-1",
        analyst_db=analyst_db,  # type: ignore[arg-type]
    )
    chunks: list[bytes] = []
    async for chunk in response.body_iterator:
        chunks.append(chunk if isinstance(chunk, bytes) else chunk.encode("utf-8"))
    body = b"".join(chunks)
    workbook = load_workbook(io.BytesIO(body), read_only=True)

    assert "Data" in workbook.sheetnames
    data_sheet = workbook["Data"]
    assert data_sheet["A1"].value == "amount"
    assert data_sheet["B1"].value == "label"
    assert data_sheet["A2"].value == 10
    assert data_sheet["B2"].value == "A"


async def _assert_chat_excel_export_writes_problematic_plotly_trace_sheet(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from core.routers.chats import save_chat_messages

    def write_tiny_png(
        self: go.Figure,
        file: str | Path,
        *_args: object,
        **_kwargs: object,
    ) -> None:
        png_data = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
        )
        Path(file).write_bytes(png_data)

    monkeypatch.setattr(go.Figure, "write_image", write_tiny_png)

    response = await save_chat_messages(
        request=None,  # type: ignore[arg-type]
        chat_id="chat-1",
        analyst_db=_FakeChartAnalystDB(),  # type: ignore[arg-type]
    )
    chunks: list[bytes] = []
    async for chunk in response.body_iterator:
        chunks.append(chunk if isinstance(chunk, bytes) else chunk.encode("utf-8"))
    if response.background:
        await response.background()
    body = b"".join(chunks)
    workbook = load_workbook(io.BytesIO(body), read_only=True)

    assert "Chart 1" in workbook.sheetnames
    chart_sheet = workbook["Chart 1"]
    assert chart_sheet["A1"].value == "x"
    assert chart_sheet["B1"].value == "y"
    assert chart_sheet["A2"].value == "FY2025"
    assert chart_sheet["B2"].value == 100
    assert chart_sheet["A3"].value == "FY2026"
    assert chart_sheet["A1"].value != "Chart Processing Error"


class _FakeDatasetHandler:
    async def get_dataframe(
        self,
        dataset_id: str,
        expected_type: DatasetType,
        max_rows: int | None = None,
    ) -> pd.DataFrame:
        assert dataset_id == "dataset-123456789"
        assert expected_type is DatasetType.ANALYST_RESULT_DATASET
        assert max_rows is None
        return pd.DataFrame({"amount": [10, 20], "label": ["A", "B"]})

    async def get_dataset_metadata(self, dataset_id: str) -> DatasetMetadata:
        assert dataset_id == "dataset-123456789"
        return DatasetMetadata(
            name=dataset_id,
            external_id=dataset_id,
            dataset_type=DatasetType.ANALYST_RESULT_DATASET,
            original_name="analysis_result",
            created_at=datetime(2026, 6, 28, tzinfo=timezone.utc),
            columns=["amount", "label"],
            original_column_types=None,
            row_count=2,
            data_source=InternalDataSourceType.GENERATED,
        )


class _FakeAnalystDB:
    def __init__(self) -> None:
        self.dataset_handler = _FakeDatasetHandler()

    async def get_chat_messages(self, chat_id: str) -> list[AnalystChatMessage]:
        assert chat_id == "chat-1"
        return [
            AnalystChatMessage(
                id="user-1",
                role="user",
                content="sum amount",
                components=[],
                in_progress=False,
            ),
            AnalystChatMessage(
                id="assistant-1",
                role="assistant",
                content="analysis",
                components=[
                    RunAnalysisResult(
                        status="success",
                        dataset_id="dataset-123456789",
                        metadata=RunAnalysisResultMetadata(
                            duration=0,
                            attempts=1,
                            datasets_analyzed=1,
                        ),
                    )
                ],
                in_progress=False,
            ),
        ]


class _FakeChartAnalystDB:
    async def get_chat_messages(self, chat_id: str) -> list[AnalystChatMessage]:
        assert chat_id == "chat-1"
        fig_json = json.dumps(
            {
                "data": [
                    {
                        "type": "scatter",
                        "x": ["FY2025", "FY2026"],
                        "y": [100],
                        "line": {"color": "red"},
                    }
                ],
                "layout": {},
            }
        )
        return [
            AnalystChatMessage(
                id="user-1",
                role="user",
                content="plot revenue",
                components=[],
                in_progress=False,
            ),
            AnalystChatMessage(
                id="assistant-1",
                role="assistant",
                content="chart",
                components=[
                    RunChartsResult(
                        status="success",
                        fig1_json=fig_json,
                        metadata=RunAnalysisResultMetadata(duration=0, attempts=1),
                    )
                ],
                in_progress=False,
            ),
        ]
